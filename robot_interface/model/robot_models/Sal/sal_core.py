
import os
import time
import rpa as r
import openpyxl
import pandas as pd


# =======================================================================
# Class to parser and handle Accesses Sheet
# =======================================================================


class Content:


    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.platform_access = []
        self.access_id = []
        self.aggregator_link = []
        self.ticket_link = []
        self.partner_id = []

        self.selected_accesses = []

        # To collect Cell info from Columns
        self.empty_platform_access = []
        self.empty_access_id = []
        self.empty_aggregator_link = []
        self.empty_ticket_link_careem = []
        self.empty_partner_id_careem = []



        # Collect Execution time

        self.tab_log = float()  # Tabalat log in
        self.tab_ord = float()  # Tabalat extract orders
        self.tab_rep = float()  # Tabalat extract reports

        self.del_log = float()  # Deliveroo log in
        self.del_ord = float()  # Deliveroo extract orders
        self.del_rep = float()  # Deliveroo extract reports
        self.del_rect_log = float()  # Deliverect  log in
        self.del_rect_ord = float()  # Deliverect  extract orders
        self.del_rect_rep = float()  # Deliverect  extract reports

        # Tabalat log in Platform: pass to 'enter talabat' function
        self.username = None
        self.password = None


class Accesses(Content):


    def __init__(self, **kwargs):
        Content.__init__(self, **kwargs)

        # Path of the Accesses Spreadsheet
        # self.filename = r"D:\Arquivos HD\Projetos HD\SD Labs\JOBS\Ahmd\rocket\rocket_kitchens\assets\Accesses sheet.xlsx"
        #
        # # Import your dataset, for example:
        # self.wb = openpyxl.load_workbook(self.filename)
        # # worksheet active:
        # self.ws = self.wb.active
        # # sheet active
        # self.sheet = self.wb["Sheet1"]


# =======================================================================
# Class to automate the leo tasks
# =======================================================================


class TaskAutomator(Accesses):

    def __init__(self, **kwargs):
        Accesses.__init__(self, **kwargs)

        # self.init = r.init(visual_automation=True)
        # self.url = r.url("https://app.careemnow.com/auth/login")

    # =====TALABAT=====================================================

    def opinion_popup_talabat(self):

        if r.read("Your opinion matters to usWe would ❤ to hear your feedback."):
            r.click(r"//button[normalize-space()='Later']")

    # Loggin Tabalat

    def enter_talabat(self, username=None,
                      password=None):

        # Get the username and password


        #OBS: get the "rate us pop up" that sometimes appear after logged.

        url = r"https://talabat.portal.restaurant/login?redirect=/"
        # get the start time
        st = time.time()

        # minimize dashboard
        r.wait(1)
        r.init(visual_automation=True)
        r.keyboard("[alt][space]")
        r.keyboard("n")
        r.wait(1)
        r.close()
        r.wait(1)
        # main task
        r.init(visual_automation=True)

        r.run("focus(title='My Restaurant')")
        r.run("maximize (title='My Restaurant')")
        r.url(url)
        r.wait(1)


        r.run("focus(title='My Restaurant')")
        r.run("maximize (title='My Restaurant')")

        r.wait(1)
        r.run("maximize (title='My Restaurant')")
        print("maximize (title='My Restaurant')")
        r.wait(1)
        print("maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")
        r.wait(2)
        print("maximizing Con")
        r.keyboard("[alt][space]")
        r.keyboard("x")
        #Verify if exist username in input field
        # Type the username and password into input field

        r.wait(2)


        r.type("//input[@id='login-email-field']", '[clear]')
        r.wait(2)
        r.click("//input[@id='login-email-field']")
        r.wait(2)

        r.type("//input[@id='login-email-field']", "{}".format('[clear]' + username))
        r.type("//input[@id='login-password-field']", "{}".format(password))
        # r.type("//input[@id='login-email-field']", "[clear]Dinarmohd@gmail.com")
        # r.type("//input[@id='login-password-field']", "MEATLABtalabat$&@")
        r.wait(1)
        r.click("//button[@id='button_login']")

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        print('Log in Tabalat execution time:', elapsed_time, 'seconds')

        self.tab_log = elapsed_time

    # Loggout Tabalat
    def exit_talabat(self):
        # get the start time
        st = time.time()

        r.click("//div[normalize-space()='Logout']")
        r.wait(2)


        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        print('Log out Tabalat execution time:', elapsed_time, 'seconds')

        self.tab_log = elapsed_time

    # Talabat Extract Reports
    def tabalat_extract_reports(self):
        # get the start time
        st = time.time()

        # main task

        r.click("//a[@data-testid='link-reports']")
        r.click("//div[@data-testid='header-wrapper']//button[1]")
        r.wait(2)
        r.click("//div[normalize-space()='last 7 days']")  # 7 days reports
        r.click("//button[normalize-space()='Submit']")  # submit
        r.wait(2)
        r.click("//div[@data-testid='performance_report-subtitle']//div//button[@type='button'][normalize-space()='Download Report']")
        r.click("//input[@value='XLSX']")  # Choose XLSX File
        r.click("//button[normalize-space()='Export Report']")  # Export Report Button

        # Download box
        r.wait(2)

        # Select the folder location and save
        r.keyboard("[enter]")

        # To choose date from range
        #r.click("//button[@aria-label='calendar view is open, go to text input view']")

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        print('Extracting Reports in Tabalat execution time:', elapsed_time, 'seconds')
        print(type(elapsed_time))
        self.tab_rep = elapsed_time

    def tab_time(self):
        time = self.tab_ord + self.tab_log + self.tab_rep
        print("Total Tabalat excution time: {} seconds".format(time))

    @staticmethod
    def talabat_close_page():

        r.keyboard("[alt][F4]")
        r.wait(1)

    def talabat_sales_per_menu_item(self):
        """
            Ranking of which menu items are the most and least popular.
        Use this to see which of your menu items are trending up or down over time.
        Note: this report does not include menu item additions like extra toppings.
        """

        # get the start time
        st = time.time()

        # main task

        r.click("//a[@data-testid='link-reports']")
        r.wait(.5)
        r.click("div[data-testid='popular_dishes-subtitle'] button[type='button']")
        r.wait(.5)
        r.click("//span[normalize-space()='Export in Excel']")
        r.wait(.5)
        r.click("//button[normalize-space()='Export Report']")

        # get the end timeb
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        print('Extracting Sales Per Menu Item \n in Tabalat execution time:', elapsed_time, 'seconds')
        print(type(elapsed_time))
        self.tab_rep = elapsed_time

    def talabat_sales_per_area(self):
        # get the start time
        st = time.time()

        # main task

        r.click("//a[@data-testid='link-reports']")

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        print('Extracting Sales by Area in Tabalat execution time:', elapsed_time, 'seconds')
        print(type(elapsed_time))
        self.tab_rep = elapsed_time



# =======================================================================
# Class to handler the orders and reports downloaded sheets
# =======================================================================


class HandlerSheet(TaskAutomator):
    # 17 = Reports
    # 30 = Finances
    # 37 = Orders

    def __init__(self):
        TaskAutomator.__init__(self)

    def talabat_read_menu_item(self):
        """
            Read menu items and let the worksheet active

        """

        # Read the latest popular dishes file:
        path = os.path.dirname('../Leo/')
        list_of_files = []

        for f_name in os.listdir(path):
            if f_name.startswith('popularDishes') and f_name.endswith('.xlsx'):
                list_of_files.append(f_name)
                print(list_of_files)
        latest_file = max(list_of_files, key=os.path.getctime)

        print("popularDishes: ", latest_file)

        # Create the PATH FOR THE order_per_day file

        self.popular_dishes = r"{}".format(str(latest_file))
        print("popularDishes: ", self.popular_dishes)

        # # Import your dataset (order_per_day), for example:
        self.wb_popular_dishes = openpyxl.load_workbook(self.popular_dishes, keep_vba=False, data_only=False)

        # order_per_day worksheet active:
        self.ws_popular_dishes = self.wb_popular_dishes.active

        # sheet active
        self.sheet_ws_popular_dishes = self.wb_popular_dishes["Sheet1"]

    def talabat_menu_item_params(self):

        self.ws = self.sheet_ws_popular_dishes

        dish = []
        total = []
        sales = []
        for row in self.ws.iter_rows(values_only=True):
            if row[0] == "Dish":
                continue
            dish.append(row[0])
            total.append(row[1])
            sales.append(row[2])
        # print('\ndish:', dish, '\ntotal:', total,  '\nsales:',  sales)
        return dish, total, sales












