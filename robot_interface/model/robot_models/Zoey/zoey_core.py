
import os
import time
import pyautogui as ptg
import logging
# =====================================================================
#  The libraries below are to handle rpa and xlsx.
# =====================================================================

import rpa as r
import openpyxl
import pandas as pd
# from rocket_kitchens_local_bot.robot_interface.model.foreground_model import get_page_title, activate_window
from foreground_model import get_page_title, activate_window
# =======================================================================
# Class to parser and handle Accesses Sheet
# =======================================================================


class Content:


    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.platform_access = []
        self.access_id = []
        self.aggregator_link = []
        self.ticket_link = []
        self.partner_id = []

        self.selected_accesses = []

        # To collect Cell info from Columns
        self.empty_platform_access = []
        self.empty_access_id = []
        self.empty_aggregator_link = []
        self.empty_ticket_link_careem = []
        self.empty_partner_id_careem = []



        # Collect Execution time

        self.tab_log = float()  # Tabalat log in
        self.tab_ord = float()  # Tabalat extract orders
        self.tab_rep = float()  # Tabalat extract reports

        self.del_log = float()  # Deliveroo log in
        self.del_ord = float()  # Deliveroo extract orders
        self.del_rep = float()  # Deliveroo extract reports
        self.del_rect_log = float()  # Deliverect  log in
        self.del_rect_ord = float()  # Deliverect  extract orders
        self.del_rect_rep = float()  # Deliverect  extract reports

        # Tabalat log in Platform: pass to 'enter talabat' function
        self.username = None
        self.password = None

class Accesses(Content):


    def __init__(self, **kwargs):
        Content.__init__(self, **kwargs)

        # Path of the Accesses Spreadsheet
        # self.filename = r"D:\Arquivos HD\Projetos HD\SD Labs\JOBS\Ahmd\rocket\rocket_kitchens\assets\Accesses sheet.xlsx"
        #
        # # Import your dataset, for example:
        # self.wb = openpyxl.load_workbook(self.filename)
        # # worksheet active:
        # self.ws = self.wb.active
        # # sheet active
        # self.sheet = self.wb["Sheet1"]

        # example use of Cell Class
        # self.cell = Cell(worksheet=self.ws, row=5, column=3)

    def lines(self):

        '''for row in self.ws.iter_rows(values_only=True, max_col=6):
            print(row)'''

        # for i in range(1, self.sheet.max_row+1):
        #     for row in self.sheet[i]:
        #         print(row.value)

        # for i in range(1, self.sheet.max_row+1):
        #     for row in self.sheet[i]:
        #         print(row.value)

        # for i in range(1, self.sheet.max_row+1):
        #     for j in range(1, self.sheet.max_column+1):
        #         print(self.sheet.cell(i, j).value)

        pass

    @staticmethod
    def cell_empty(val=None, n=None) -> True:
        val = n.value
        if val is None:
            return True

        ...

    def columns(self) -> None:
        ...

    def platform(self) -> None:
        '''
            This class shows the Platform Access.
            Column A

        :return: None

        '''

        for i in self.sheet['A2':'A25']:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.info(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_platform_access.append(coor)

        logging.info("Empty cells: ", self.empty_platform_access)

    def username(self):
        '''
            This class shows the Access ID.
            Column B

        :return: None

        '''

        # for col in self.ws.iter_cols(values_only=True, max_col=6):
        #     logging.info(col)
        username_column = self.sheet['B2':'B25']
        for i in username_column:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.info(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_access_id.append(row_number)

        current_username = []
        for i in range(1, 22):
            ...

        logging.info("Size Username: {}".format(len(username_column)))
        logging.info("Empty cells: ", self.empty_access_id)
        pass

    def credentials(self):
        '''
            This class shows the Access ID.
            Column C

        :return: None

        '''

        # for col in self.ws.iter_cols(values_only=True, max_col=6):
        #     print(col)

        for i in self.sheet['C1':'C25']:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.info(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_access_id.append(row_number)

        logging.info("Empty cells: ", self.empty_access_id)
        pass

    def links(self):
        '''
            This class shows the Aggregator Link.
            Column D

        :return: None

        '''

        for i in self.sheet['D1':'D25']:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.info(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_aggregator_link.append(row_number)

        logging.info("Empty cells: ", self.empty_aggregator_link)
        pass

    def ticket_link_careem(self):
        '''
            This class shows the Ticket_link (Careem).
            Column E

        :return: None

        '''

        # for col in self.ws.iter_cols(values_only=True, max_col=6):
        #     logging.info(col)

        for i in self.sheet['E1':'E25']:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.info(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_ticket_link_careem.append(row_number)

        logging.info("Empty cells: ", self.empty_ticket_link_careem)

    def partner_id_careem(self):
        '''
            This class shows the Partner ID (Careem).
            Column F

        :return: None

        '''

        # for col in self.ws.iter_cols(values_only=True, max_col=6):
        #     logging.info(col)

        for i in self.sheet['F1':'F25']:
            for n in i:  # read the cell in sheet
                coor = n.coordinate
                row_number = n.row
                val = n.value
                logging.infoprint(coor, row_number, val)

                if self.cell_empty(val, n):
                    self.empty_partner_id_careem.append(row_number)

        logging.info("Empty cells: ", self.empty_partner_id_careem)

    def select_cell(self):

        # self.sheet['{}{}'.format("c", "3")].value

        pass

    def max_rows(self):
        self.max_col = self.sheet.max_column
        logging.info(self.sheet.max_row)
        return self.max_col

    def max_columns(self):
        logging.info(self.sheet.max_column)
        self.max_col = self.sheet.max_column
        return self.max_col


# =======================================================================
# Class to automate the tasks
# =======================================================================


class TaskAutomator(Accesses):

    def __init__(self, **kwargs):
        Accesses.__init__(self, **kwargs)

        # self.init = r.init(visual_automation=True)
        # self.url = r.url("https://app.careemnow.com/auth/login")

    def use_another_email(self):
        pass

    def user(self, name=None, password=None):
        if name and password is None:
            name = 'ziryabalquoz@careemnow.com'
            password = 'cn.ziryabalquoz'

            return name, password

    # =====CAREEM==================================================================

    # Loggin Careem
    @staticmethod
    def enter_careem():
        # get the start time
        st = time.time()


        # main task
        r.init(visual_automation=True)
        r.run("maximize  (title='My Restaurant)")
        r.url("https://app.careemnow.com/auth/login")

        logging.info("maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")


        try:

            log_another = r.exist("//span[contains(text(),'Log in with another email')]")
            logging.info(log_another)

            r.wait(1)
            # If Log in with another email button don't exist: pass
            if log_another is False:
                pass


            r.wait(1)
            # Click in Log with another email if exists
            r.click("//span[contains(text(),'Log in with another email')]")
            r.wait(1)

        except:
            pass

        r.type("//div[@class='ant-form-item-control-input']", 'ziryabalquoz@careemnow.com')
        r.click("//button[@type='submit']")
        r.wait(1)
        r.type("//div[@class='ant-form-item-control-input-content']", "cn.ziryabalquoz")
        r.wait(1)
        r.click("//button[@type='submit']")

        r.wait(1)
        r.close()

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Log in Careem execution time:', elapsed_time, 'seconds')

    # Loggout Careem
    @staticmethod
    def exit_careem():
        # get the start time
        st = time.time()

        # main task
        r.init(visual_automation=True)
        r.run("maximize (title='My Restaurant)")
        r.url("https://app.careemnow.com/auth/login")

        r.wait(1)
        logging.info("maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")
        logging.info("logging out")

        # r.wait(1)
        # Click in center to close google translator pop-up
        # r.click("//button[2]")

        r.wait(1)
        # Clink in navbar to get the loggout button
        r.click("//li[@class='ant-menu-submenu ant-menu-submenu-horizontal']")
        r.click("Ziryab Al Quoz 2 Call center")
        r.click("Log Out")
        # r.click("Log out")
        r.wait(1)
        r.close()
        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Log out Careem execution time:', elapsed_time, 'seconds')

    def careem_orders(self):
        ...

    def careem_past_orders(self):
        ...


    # =====DELIVEROO==================================================================

    # Loggin Deliveroo
    def enter_deliveroo(self):
        url = r"https://restaurant-hub.deliveroo.net/login?redirect=/analytics"
        # get the start time
        st = time.time()

        # main task
        r.init(visual_automation=True)
        r.run("maximize (title='My Restaurant)")
        r.url(url)

        r.wait(1)
        logging.info("maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")

        # Verify if exist username in input field
        # Type the username and password into input field

        r.wait(2)

        r.type("//input[@placeholder='e.g. Joe.Bloggs@deliveroo.com']", '[clear]')
        r.wait(2)
        r.click("//input[@placeholder='e.g. Joe.Bloggs@deliveroo.com']")
        r.wait(2)
        r.type("//input[@placeholder='e.g. Joe.Bloggs@deliveroo.com']", "[clear]ahmd@ziryab.io")
        r.type("//input[@placeholder='e.g. ********']", "YYf898zH1!")
        r.wait(1)
        r.click("//button[@aria-label='Log in']")

        r.wait(5)
        # r.close()

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Log in Deliveroo execution time:', elapsed_time, 'seconds')

        self.del_log = elapsed_time
        pass

    # Loggout Careem
    def exit_deliveroo(self):
        pass

    # =====DELIVERECT==================================================================

    # Loggin Deliverect
    def enter_deliverect(self):
        url = r"https://www.deliverect.com/en"
        # get the start time
        st = time.time()

        # main task
        # os.system("pkill chrome")
        ptg.hotkey('ctrl', 'w')

        r.init(visual_automation=True)
        r.run("maximize (title='My Restaurant)")
        r.url(url)

        logging.info("maximizing")
        r.wait(3)


        r.keyboard("[alt][space]")
        r.keyboard("x")

        # Verify if exist username in input field
        # Type the username and password into input field

        r.wait(2)
        # Click in DIV {Sign in}
        r.click("//body//div//div//div//div//header//div//div//div//a[normalize-space()='Sign in']")
        # r.wait(8)

        logging.info("get screen size")

        logging.info(ptg.size())
        logging.info("get mouse position")

        logging.info(ptg.position())

        r.read("//input[@placeholder='yours@example.com']")
        r.type("//input[@placeholder='yours@example.com']", '[clear]')

        r.wait(2)
        r.type("//input[@placeholder='yours@example.com']", "[clear]bots@rocketkitchens.com")
        r.type("//input[@id='1-password']", "<Hern8=E")
        r.wait(1)
        r.click("//button[@aria-label='Log In']")

        r.wait(15)
        r.close()

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Log in Deliverect execution time:', elapsed_time, 'seconds')

        self.del_rect_log = elapsed_time
        pass

    # Loggout Deliverect
    def exit_deliverect(self):
        pass

    # =====TALABAT=====================================================

    def opinion_popup_talabat(self):

        if r.read("Your opinion matters to usWe would ❤ to hear your feedback."):
            r.click(r"//button[normalize-space()='Later']")

    # Loggin Tabalat

    def enter_talabat(self, username=None, password=None):
        # Get the username and password

        # OBS: get the "rate us pop up" that sometimes appear after logged.

        url = r"https://talabat.portal.restaurant/login?redirect=/"

        # Configure logging
        logger = logging.getLogger(__name__)

        # get the start time
        st = time.time()

        # variables

        # # minimize dashboard
        # r.wait(1)
        # r.init(visual_automation=True)
        # r.keyboard("[alt][space]")
        # r.keyboard("n")
        # r.wait(1)
        # r.close()
        # r.wait(1)
        # main task

        r.init(visual_automation=True)

        title = get_page_title("https://talabat.portal.restaurant/login?redirect=/")
        focus = "focus(title='{}')".format(title)
        maximize = "maximize (title='{}')".format(title)
        r.run(focus)
        r.run(maximize)
        r.url(url)
        r.wait(1)

        r.run(focus)
        r.run(maximize)
        activate_window(title)
        r.wait(1)
        r.run(maximize)
        logger.info("Maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")
        r.wait(2)
        logger.info("Maximizing Con")
        r.keyboard("[alt][space]")
        r.keyboard("x")

        # Verify if exist username in input field
        # Type the username and password into input field

        r.wait(2)

        r.type("//input[@id='login-email-field']", '[clear]')
        r.wait(2)
        r.click("//input[@id='login-email-field']")
        r.wait(2)

        r.type("//input[@id='login-email-field']", "{}".format('[clear]' + username))
        r.type("//input[@id='login-password-field']", "{}".format(password))
        # r.type("//input[@id='login-email-field']", "[clear]Dinarmohd@gmail.com")
        # r.type("//input[@id='login-password-field']", "MEATLABtalabat$&@")
        r.wait(1)
        r.click("//button[@id='button_login']")
        r.wait(3)
        title = get_page_title("https://talabat.portal.restaurant/dashboard")
        activate_window(title)

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logger.info('Log in Tabalat execution time: %s seconds', elapsed_time)

        self.tab_log = elapsed_time

    # Loggout Tabalat
    def exit_talabat(self):
        # get the start time
        st = time.time()

        r.click("//div[normalize-space()='Logout']")
        r.wait(2)

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st

        # Log execution time
        logging.info('Log out Tabalat execution time: %s seconds', elapsed_time)

        self.tab_log = elapsed_time

    # =====ZOMATO=====================================================

    # Loggin Zomato
    def enter_zomato(self):
        url = r"https://www.zomato.com/partners/login"
        # get the start time
        st = time.time()

        # main task
        r.init(visual_automation=True)
        r.run("maximize (title='My Restaurant)")
        r.url(url)

        logging.info("maximizing")
        r.keyboard("[alt][space]")
        r.keyboard("x")

        r.wait(1)

        r.click("//span[contains(text(),'Login')]")
        r.wait(3)
        # r.click("//span[normalize-space()='Continue with Email']")
        r.click("Continue with Email")
        r.wait(1)
        r.type("Email", 'ahmd@rocketkitchens.com')

        r.wait(20)
        r.close()

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Log in Zomato execution time:', elapsed_time, 'seconds')
        pass

    # Loggout Zomato
    def exit_zomato(self):
        pass

    # Talabat Extract Orders
    def tabalat_extract_orders(self):
        # get the start time
        st = time.time()

        # main task
        r.click("//a[@data-testid='link-orders']")  # Today
        r.wait(2)
        r.click("//span[normalize-space()='Today']")

        logging.info('Checking no_orders')
        no_orders = r.get_text("No Orders to show")
        logging.info(f'no_orders value: {no_orders}')
        if no_orders == 'No Orders to show':
            r.click("//span[normalize-space()='Yesterday']")
            r.dclick("//span[normalize-space()='Yesterday']")

        r.wait(1)
        r.click("//div[@class='plugin-muiv4-MuiBox-root plugin-muiv4-jss6']//button[@type='button']")  # Export Button
        # Export Orders Details
        r.click("//div[normalize-space()='Today']")  # Today
        r.click("//input[@value='XLSX']")  # xlsx
        r.click("//span[normalize-space()='Download Report']")

        r.wait(2)

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info(f'Extracting Orders in Tabalat execution time: {elapsed_time} seconds')
        logging.info(f'elapsed_time type: {type(elapsed_time)}')
        self.tab_ord = elapsed_time

    def last_week_orders(self):
        # create logger
        logger = logging.getLogger(__name__)

        # get the start time
        st = time.time()

        # main task
        logger.info("clicking in orders")
        r.wait(2)
        r.click("//a[@data-testid='link-orders']")
        r.click("//a[@data-testid='link-orders']")
        r.click("//a[@data-testid='link-orders']")
        r.wait(2)
        r.click("//a[@data-testid='link-orders']")
        logger.info("orders clicked ")

        # Last Week
        r.click("//body//div//div[@data-testid='content-container-new']//div//div//div//div//div[2]//div[1]//button[1]")
        r.wait(2)
        r.click("//div[normalize-space()='Last week']")

        r.wait(2)
        # Select .xlsx format to download
        r.click("//input[@value='XLSX']")  # xlsx
        r.wait(2)
        # Download Report
        r.click("//span[normalize-space()='Download Report']")

        r.wait(2)

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logger.info('Extracting Orders in Tabalat execution time: %s seconds', elapsed_time)
        # logging.info(type(elapsed_time))
        self.tab_ord = elapsed_time

    # Talabat Extract Reports
    def tabalat_extract_reports(self):
        # get the start time
        st = time.time()

        # main task

        r.click("//a[@data-testid='link-reports']")
        r.click("//div[@data-testid='header-wrapper']//button[1]")
        r.wait(2)
        r.click("//div[normalize-space()='last 7 days']")  # 7 days reports
        r.click("//button[normalize-space()='Submit']")  # submit
        r.wait(2)
        r.click \
            ("//div[@data-testid='performance_report-subtitle']//div//button[@type='button'][normalize-space()='Download Report']")
        r.click("//input[@value='XLSX']")  # Choose XLSX File
        r.click("//button[normalize-space()='Export Report']")  # Export Report Button

        # Download box
        r.wait(2)

        # Select the folder location and save
        r.keyboard("[enter]")

        # To choose date from range
        # r.click("//button[@aria-label='calendar view is open, go to text input view']")

        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Extracting Reports in Tabalat execution time:', elapsed_time, 'seconds')
        logging.info(type(elapsed_time))
        self.tab_rep = elapsed_time

    def tab_time(self):
        time = self.tab_ord + self.tab_log + self.tab_rep
        logging.info("Total Tabalat excution time: {} seconds".format(time))

    def talabat_close_page(self):

        r.keyboard("[alt][F4]")
        r.wait(1)


    # Deliveroo Extract Orders
    def deliveroo_extract_orders(self):
        # get the start time
        st = time.time()

        # main task

        r.click("//a[@href='/orders']//div//p")  # Sales
        r.wait(4)
        r.click("//i[normalize-space()='save_alt']")  # Export orders data


        r.wait(2)
        r.click("//input[@id='All organisations and sites']")  # All Organizationand sites
        # Export Orders Details
        r.click("//button[@aria-label='Export'] ")  # Export
        r.wait(2)
        r.click("//button[@aria-label='Close']")    # Close

        logging.info("We’re creating a report and it’ll be \n in your email inbox in 5-10 minutes.")


        r.wait(2)


        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Extracting Orders in Deliveroo execution time:', elapsed_time, 'seconds')

        self.del_ord = elapsed_time

    # Deliveroo Total Time Execution
    def del_time(self):
        time = self.del_rect_ord + self.del_rect_log + self.del_rect_rep
        logging.info("Total Deliverect execution time: {} seconds".format(time))

    # Deliverect Extract Orders
    def deliverect_extract_orders(self):
        # get the start time
        st = time.time()

        # main task

        r.click("//a[@href='/orders']//div//p")  # Sales
        r.wait(4)
        r.click("//i[normalize-space()='save_alt']")  # Export orders data


        ptg.position()


        r.wait(2)
        r.click("//input[@id='All organisations and sites']")  # All Organizationand sites
        # Export Orders Details
        r.click("//button[@aria-label='Export'] ")  # Export
        r.wait(2)
        r.click("//button[@aria-label='Close']")    # Close

        logging.info("We’re creating a report and it’ll be \n in your email inbox in 5-10 minutes.")


        r.wait(2)


        # get the end time
        et = time.time()

        # get the execution time
        elapsed_time = et - st
        logging.info('Extracting Orders in Deliveroo execution time:', elapsed_time, 'seconds')

        self.del_ord = elapsed_time

    # Deliverect Total Time Execution
    def del_rect_time(self):
        time = self.del_ord + self.del_log + self.del_rep
        logging.info("Total Deliveroo excution time: {} seconds".format(time))


# =======================================================================
# Class to handler the orders and reports downloaded sheets
# =======================================================================


class HandlerSheet(TaskAutomator):
    # 17 = Reports
    # 30 = Finances
    # 37 = Orders

    def __init__(self):
        TaskAutomator.__init__(self)
        self.logger = logging.getLogger(__name__)

    # ======================================
    # Talabat Orders
    # ======================================

    # Zoey Bot:
    def tabalat_process_orders(self):
        pass

    # Read Order Details - is it Zoey too
    def talabat_read_orders(self):
        """
        Read order details and let the worksheet active
        """

        # Path of the Accesses Spreadsheet

        self.filename = r"../LocalBot/orderDetails.xlsx"

        # Import your dataset, for example:
        self.wb = openpyxl.load_workbook(self.filename, data_only=False)
        # worksheet active:
        self.ws = self.wb.active
        # sheet active
        self.sheet = self.wb["Sheet1"]

    # Read (after downloaded) the latest ordersPerDay and active worksheet
    def talabat_read_average_order_per_day(self):
        # Set up logging
        logging.basicConfig(filename='example.log', level=logging.INFO)

        # Read the latest ordersPerDay file:
        path = os.path.dirname('../LocalBot/')
        list_of_files = []
        for f_name in os.listdir(path):
            if f_name.startswith('ordersPerDay') and f_name.endswith('.xlsx'):
                list_of_files.append(f_name)
                logging.info(f"List of files: {list_of_files}")
        latest_file = max(list_of_files, key=os.path.getctime)
        logging.info(f"Latest file: {latest_file}")

        # Create the PATH FOR THE order_per_day file
        self.order_per_day = r"{}".format(str(latest_file))
        logging.info(f"Order per day: {self.order_per_day}")

        # # Import your dataset (order_per_day), for example:
        self.wb_order_per_day = openpyxl.load_workbook(self.order_per_day, keep_vba=False, data_only=False)
        # order_per_day worksheet active:
        self.ws_opd = self.wb_order_per_day.active
        # sheet active
        self.sheet_ws_opd = self.wb_order_per_day["Sheet1"]

    # Using the anterior worksheet ordersPerDay
    # Get the average order per day from column B
    def tabalat_average_order_per_day(self):
        """
        Sum all Food Value Column
        -------------------------

        Go to Orders Column to find
        average order per day
        """

        max_row = "B{}".format(self.sheet_ws_opd.max_row)
        len_average_order = self.sheet_ws_opd.max_row + 1
        logging.info('Getting average order')

        # for food_value in self.sheet['S2':max_row]:
        #     for n in food_value:
        #         logging.info(n.value)

        logging.info("max order row: %s", max_row)
        logging.info("average value: %s", len_average_order)
        self.average = "B{}".format(len_average_order)
        logging.info("average: %s", self.average)

        self.sheet_ws_opd[self.average] = '= AVERAGE(B2:{})'.format(max_row)

        # Save the File with the average updated
        self.wb_order_per_day.save(self.order_per_day)

    def read_value_from_excel(self, filename, column="B", row=8):
        """Read a single cell value from an Excel file"""
        return pd.read_excel(filename,
                             sheet_name="Sheet1",
                             skiprows=None,
                             usecols="U",
                             nrows=None,
                             header=None,
                             names=None)


    # Solve it:
    # s = sum
    # u = sum
    # v = (sN-uN)*30%
    # w = s-u-v

    # Order Details Not Order Per Day

    # right -> =SOMA(S2:SN) where N is the last cell in row

    def tabalat_average_food(self):
        """
                            Sum all Food column

                        """
        # 19/02/23 -> Food Value Column changed from S to V column.
        # Verify how much lines each column has
        # it count maximum lines in the spreadsheet
        max_row = "V{}".format(self.sheet.max_row)

        # it get the maximum row and add plus 1 in S column
        len_discount_value = self.sheet.max_row + 1
        # len_discount_value = self.sheet.max_row
        logging.info('Sum all Food column')
        # for food_value in self.sheet['U2':max_row]:  # section to extraction of table names for the new tables.
        #     for n in food_value:
        #         logging.info(n.value)

        logging.info("max rox: ", max_row)
        logging.info("Food value: ", len_discount_value)
        average = "V{}".format(len_discount_value)

        # Save the food average value
        self.sheet[average] = '= SUM(V2:{})'.format(max_row)

        # Save the File with the average updated
        self.wb.save(self.filename)


    # Order Details Not Order Per Day

    # right -> =SOMA(U2:UN)where N is the last cell in row
    def tabalat_average_discount(self):
        """
                Sum all Discount column
                """

        # 19/02/23 -> Discount Column changed from U to X column.
        max_row = "X{}".format(self.sheet.max_row)
        len_discount_value = self.sheet.max_row + 1

        self.logger.info('Summing all Discount column')
        self.logger.info("max row: %s", max_row)
        self.logger.info("discount value: %s", len_discount_value)

        average = "X{}".format(len_discount_value)
        self.sheet[average] = '= SUM(X2:{})'.format(max_row)

        self.wb.save(self.filename)

    # right -> =S41-U42 where N is the

    # Order Details Not Order Per Day
    def tabalat_average_comission(self):
        logger = logging.getLogger(__name__)
        ws = self.sheet
        max_row_for_v = max((v.row for v in ws['V'] if v.value is not None))
        max_row_for_x = max((x.row for x in ws['X'] if x.value is not None))
        logger.info("Max Row for Column V: %s, type: %s", max_row_for_v, type(max_row_for_v))
        logger.info("Max Row for Column X: %s, type: %s", max_row_for_x, type(max_row_for_x))

        last_v_value = ws["V{}".format(max_row_for_v)].value
        last_x_value = ws["X{}".format(max_row_for_x)].value
        max_row_v = ws["V{}".format(int(max_row_for_v))].coordinate
        max_row_x = ws["X{}".format(int(max_row_for_x))].coordinate

        logger.info("last_v_value: %s", last_v_value)
        logger.info("last_x_value: %s", last_x_value)
        logger.info('max_row_v: %s', max_row_v)
        logger.info('V value: %s', ws[max_row_v])
        logger.info('X value: %s', ws[max_row_x])

        ws["AE51"] = "AVG comission"
        ws["AE50"] = '={}-{}'.format(max_row_v, max_row_x)

        logger.info("Value in cell AE50: %s", ws["AE50"].value)

        self.wb.save(self.filename)

    # Order Details Not Order Per Day
    def tabalat_gross_profit(self):
        # Configure logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

        # read the last details order
        ws = self.sheet

        # Read the latest ordersPerDay file:
        path = os.path.dirname('../LocalBot/')
        list_of_files = []
        for f_name in os.listdir(path):
            if f_name.startswith('orderDetails') and f_name.endswith('.xlsx'):
                list_of_files.append(f_name)
                logging.info("Found file: %s", f_name)
        latest_file = max(list_of_files, key=os.path.getctime)
        logging.info("Latest orderDetails file: %s", latest_file)

        # Create the PATH FOR THE order_per_day file
        self.order_details = r"{}".format(str(latest_file))
        logging.info("Order Details: %s", self.order_details)

        # Import your dataset (order_per_day), for example:
        self.wb_order_details = openpyxl.load_workbook(self.order_details, data_only=False)

        # order_details worksheet active:
        self.ws = self.wb_order_details.active
        logging.info("Cell values: %s, %s, %s", self.wb_order_details["Sheet1"]["AE50"], self.ws["AE50"].value,
                     self.ws["U94"].value)
        time.sleep(1)

        food_values = []
        max_row = "V{}".format(self.ws.max_row)
        logging.info('Getting average food value')
        for food_value in self.ws['V2':max_row]:
            for n in food_value:
                if n.value is not None:
                    food_values.append(n.value)
        time.sleep(1)

        logging.info('Food values: %s', food_values[:-1])
        sum_food_values = sum(food_values[:-1])
        logging.info('Sum food values: %s', sum_food_values)
        time.sleep(1)

        discount_values = []
        max_row = "X{}".format(self.ws.max_row)
        logging.info('Getting average discount')
        for discount in self.ws['X2':max_row]:
            for n in discount:
                if n.value is not None:
                    discount_values.append(n.value)

        time.sleep(1)
        logging.info('Discount values: %s', discount_values[:-1])
        sum_discount_values = sum(discount_values[:-1])
        logging.info('Sum discount values: %s', sum_discount_values)

        # Comission
        # s-u
        avg_comission = (sum_food_values - sum_discount_values) * (30 / 100)
        logging.info('Avg comission: %s', avg_comission)

        # Gross Profit
        gross_profit = sum_food_values - sum_discount_values - avg_comission
        logging.info('Gross profit: %s', gross_profit)

        # Save the file with the average updated
        self.wb.save(self.filename)

        return sum_discount_values, sum_food_values, avg_comission, gross_profit

    def tabalat_get_average_order_value(self):
        self.wb_order_per_day = openpyxl.load_workbook(self.order_per_day, data_only=True)
        self.sheet_ws_opd = self.wb_order_per_day["Sheet1"]
        avg_coord = self.sheet_ws_opd[self.average].coordinate
        avg_value = self.sheet_ws_opd[self.average].value
        logging.info("avg_coord :",  avg_coord)
        # logging.info(avg_value)
        x = self.sheet_ws_opd["{}".format(str(avg_coord))]



        logging.info(x)
        logging.info("avg_coord: ", x)
        # logging.info("avg_value: " , self.sheet_ws_opd[avg_value])

        ...

    # Get the last Order Details and turn in to a path
    # Call and Active the Worksheet
    def talabat_order_details_filename(self):
        """
            -> Path of the the last Order
            -> Details Spreadsheet
            ================================================
            -> self.filename = r"../LocalBot/orderDetails.xlsx"
        """


        path = os.path.dirname('../LocalBot/')
        list_of_files = []
        for f_name in os.listdir(path):
            if f_name.startswith('orderDetails') and f_name.endswith('.xlsx'):
                list_of_files.append(f_name)
                logging.info(list_of_files)
        latest_file = max(list_of_files, key=os.path.getctime)
        logging.info("latest_file orderDetails: ", latest_file)

        self.order_details = r"{}".format(str(latest_file))
        logging.info("orderDetails': ", self.order_details)

        # # Import your dataset, for example:
        self.wb_order_details = openpyxl.load_workbook(self.order_details, data_only=True)
        # worksheet active:
        self.ws_od = self.wb_order_details.active
        # sheet active
        self.sheet_ws_od = self.wb_order_details["Sheet1"]

        return self.wb_order_details

    # filename = talabat_order_details_filename()

    # ======================================
    #  Talabat Reports
    # ======================================

    # Leo Bot:


    def tabalat_process_reports(self):
        pass

    def tabalat_report_sales_per_menu(self):
        pass

    def tabalat_report_sales_by_area(self):
        pass

    def delete_output_file(self, filename):
        home_dir = os.path.expanduser("~")
        output_dir = os.path.join(home_dir, "Downloads", "output")
        filepath = os.path.join(output_dir, filename)

        if os.path.exists(filepath):
            os.remove(filepath)
            logging.info(f"{filepath} deleted.")
        else:
            logging.info(f"{filepath} does not exist.")









    # ======================================
    #  Talabat Financials Reports
    # ======================================

    # ======================================
    #  Talabat Financials Reports
    # ======================================


    def deliveroo_process_orders(self):
        pass

    def deliveroo_process_orders(self):
        pass

    def deliveroo_process_reports(self):
        pass

    def careem_process_orders(self):
        pass

    # ======================================
    #
    # ======================================
